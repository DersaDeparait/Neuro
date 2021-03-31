from openpyxl import load_workbook
from openpyxl import Workbook

class ExcelManager:
    def __init__(self, filename = "excel_output/0.xlsx", letter_name= "Sheet5"):
        self.filename = filename
        self.letter_name=letter_name
        self.open_book()
        self.open_letter()
    def open_book(self):
        try:
            self.wb = load_workbook(filename = self.filename, data_only=True)
        except:
            self.wb = Workbook()
            self.wb.save(filename = self.filename)
            self.wb = load_workbook(filename=self.filename, data_only=True)
    def open_letter(self):
        try:
            self.letter = self.wb[self.letter_name]
        except:
            self.wb.create_sheet(self.letter_name)
            self.letter = self.wb[self.letter_name]
            self.save_data()

    def read(self):
        '''Зчитує коробку даних, прямокутник розміри якого рівні максимальним значенням колонок і рядків для даних комірок'''
        i = 0
        data = []
        for row in self.letter.iter_rows(min_row=1):
            i += 1
            data.append([])
            for j in range(len(row)):
                data[-1].append(row[j].value)
        return data
    def read_rect(self, min_row=1, min_col=1, max_col=3, max_row=6):
        '''Зчитує прямокутну коробку даних і повертає їх значення'''
        data = []
        for row in self.letter.iter_rows(min_row = min_row, min_col = min_col, max_col = max_col, max_row = max_row):
            data.append([])
            for cell in row:
                data[-1].append(cell.value)
        return data
    def read_one_cell(self, row = 1, column = 1):
        '''Зчитує одну конктретну колонку'''
        d = self.letter.cell(row=row, column=column)
        return d.value

    def write_one_cell(self, row = 1, column = 1, value=None):
        '''Записує одну конкретну колонку'''
        d = self.letter.cell(row = row, column = column, value = value)
        self.save_data()
        return d

    def write_append(self, weights = [45, 105, 100, 0, 50, 70]):
        self.letter.append(weights)
        self.save_data()
    def write_rect(self, start_row = 1, start_column = 1, data = [[1, 2, 3, 4, 5], [6, 7, 8, 9, 10], [11, 12, 13, 14, 15, 16, 17]]):
        for i, row in zip(range(len(data)), self.letter.iter_rows(
                                         min_row = start_row,
                                         min_col = start_column,
                                         max_col = start_column - 1 + max([len(r) for r in data]),
                                         max_row = start_row - 1 + len(data))):
            for j, cell in zip(range(len(data[i])), row):
                cell.value = data[i][j]
        self.save_data()
        return data

    def save_data(self):
        self.wb.save(self.filename)
        self.open_book()
